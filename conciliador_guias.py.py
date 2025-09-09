import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import seaborn as sns

class ConciliadorGuias:
    def __init__(self):
        self.setup_directories()
        
    def setup_directories(self):
        """Cria diretórios necessários"""
        directories = ['dados', 'relatorios', 'graficos', 'conciliacao']
        for directory in directories:
            if not os.path.exists(directory):
                os.makedirs(directory)
    
    def gerar_dados_ficticios(self):
        """Gera dados fictícios para demonstração"""
        np.random.seed(42)
        
        # Dados básicos
        convenios = ['Unimed', 'Bradesco', 'SulAmerica', 'Amil', 'NotreDame']
        procedimentos = [
            'Consulta Cardiologia', 'Exame ECG', 'Ultrassom Abdomen', 'Raio-X Torax',
            'Consulta Ortopedia', 'Ressonancia Magnetica', 'Tomografia', 'Endoscopia',
            'Hemograma Completo', 'Glicemia', 'Colesterol Total', 'Ureia e Creatinina'
        ]
        
        # 1. Sistema Interno (o que o hospital enviou)
        guias_internas = []
        for i in range(300):  # 300 guias
            numero_guia = f"GH{2024}{str(i+1).zfill(6)}"  # GH2024000001
            convenio = np.random.choice(convenios)
            
            guia = {
                'numero_guia': numero_guia,
                'convenio': convenio,
                'data_atendimento': datetime.now() - timedelta(days=np.random.randint(1, 60)),
                'paciente': f"Paciente {i+1:03d}",
                'procedimento': np.random.choice(procedimentos),
                'valor_enviado': round(np.random.uniform(50, 1500), 2),
                'status_interno': np.random.choice(['Enviado', 'Reenviado', 'Aguardando'], p=[0.8, 0.15, 0.05]),
                'data_envio': datetime.now() - timedelta(days=np.random.randint(1, 30)),
                'lote': f"LT{np.random.randint(1000, 9999)}"
            }
            guias_internas.append(guia)
        
        df_interno = pd.DataFrame(guias_internas)
        
        # 2. Demonstrativo da Operadora (o que a operadora processou)
        # Simula algumas situações reais:
        # - 70% das guias pagas corretamente
        # - 15% com valores divergentes
        # - 10% glosadas
        # - 5% não processadas (não aparecem no demonstrativo)
        
        guias_operadora = []
        
        for _, guia_interna in df_interno.iterrows():
            situacao = np.random.choice(['paga_correta', 'valor_divergente', 'glosada', 'nao_processada'], 
                                      p=[0.70, 0.15, 0.10, 0.05])
            
            if situacao == 'nao_processada':
                continue  # Não aparece no demonstrativo
            
            guia_op = {
                'numero_guia': guia_interna['numero_guia'],
                'convenio': guia_interna['convenio'],
                'data_processamento': datetime.now() - timedelta(days=np.random.randint(1, 15)),
                'paciente': guia_interna['paciente'],
                'procedimento': guia_interna['procedimento']
            }
            
            if situacao == 'paga_correta':
                guia_op['valor_pago'] = guia_interna['valor_enviado']
                guia_op['status_operadora'] = 'Pago'
                guia_op['motivo_glosa'] = ''
                
            elif situacao == 'valor_divergente':
                # Valor diferente (pode ser menor por co-participação, ou por tabela diferente)
                fator = np.random.choice([0.7, 0.8, 0.9, 1.1, 1.2], p=[0.3, 0.3, 0.2, 0.1, 0.1])
                guia_op['valor_pago'] = round(guia_interna['valor_enviado'] * fator, 2)
                guia_op['status_operadora'] = 'Pago com Divergência'
                guia_op['motivo_glosa'] = np.random.choice([
                    'Valor tabela diferente', 'Co-participação aplicada', 
                    'Desconto contratual', 'Valor não conferido'
                ])
                
            elif situacao == 'glosada':
                guia_op['valor_pago'] = 0.00
                guia_op['status_operadora'] = 'Glosado'
                guia_op['motivo_glosa'] = np.random.choice([
                    'Falta de documentação', 'Autorização inválida', 'Código incorreto',
                    'Prazo de envio vencido', 'Procedimento não coberto'
                ])
            
            guias_operadora.append(guia_op)
        
        df_operadora = pd.DataFrame(guias_operadora)
        
        # Salvar dados
        df_interno.to_csv('dados/guias_sistema_interno.csv', index=False)
        df_operadora.to_csv('dados/demonstrativo_operadora.csv', index=False)
        
        print("✅ Dados fictícios gerados:")
        print(f"   • {len(df_interno)} guias no sistema interno")
        print(f"   • {len(df_operadora)} guias no demonstrativo da operadora")
        
        return df_interno, df_operadora
    
    def carregar_dados(self):
        """Carrega dados dos arquivos CSV"""
        try:
            df_interno = pd.read_csv('dados/guias_sistema_interno.csv')
            df_interno['data_atendimento'] = pd.to_datetime(df_interno['data_atendimento'])
            df_interno['data_envio'] = pd.to_datetime(df_interno['data_envio'])
            
            df_operadora = pd.read_csv('dados/demonstrativo_operadora.csv')
            df_operadora['data_processamento'] = pd.to_datetime(df_operadora['data_processamento'])
            
            return df_interno, df_operadora
            
        except FileNotFoundError:
            print("❌ Arquivos não encontrados. Gerando dados fictícios...")
            return self.gerar_dados_ficticios()
    
    def conciliar_guias(self, df_interno, df_operadora):
        """Realiza a conciliação entre sistema interno e operadora"""
        
        # Merge das bases
        conciliacao = df_interno.merge(
            df_operadora[['numero_guia', 'valor_pago', 'status_operadora', 'motivo_glosa', 'data_processamento']], 
            on='numero_guia', 
            how='left'
        )
        
        # Preencher valores não encontrados
        conciliacao['valor_pago'] = conciliacao['valor_pago'].fillna(0)
        conciliacao['status_operadora'] = conciliacao['status_operadora'].fillna('Não Processado')
        conciliacao['motivo_glosa'] = conciliacao['motivo_glosa'].fillna('Guia não encontrada no demonstrativo')
        
        # Calcular divergências
        conciliacao['diferenca_valor'] = conciliacao['valor_pago'] - conciliacao['valor_enviado']
        conciliacao['percentual_divergencia'] = (
            (conciliacao['diferenca_valor'] / conciliacao['valor_enviado']) * 100
        ).round(2)
        
        # Classificar tipos de divergência
        def classificar_divergencia(row):
            if pd.isna(row['data_processamento']):
                return '🔴 Não Processado'
            elif row['valor_pago'] == 0:
                return '❌ Glosado'
            elif abs(row['diferenca_valor']) < 0.01:  # Diferença menor que 1 centavo
                return '✅ OK - Valores Conferem'
            elif row['diferenca_valor'] > 0:
                return '⚠️ Pago a Maior'
            else:
                return '🟡 Pago a Menor'
        
        conciliacao['classificacao'] = conciliacao.apply(classificar_divergencia, axis=1)
        
        # Adicionar dias em aberto para itens não processados
        conciliacao['dias_em_aberto'] = (
            datetime.now() - conciliacao['data_envio']
        ).dt.days
        
        return conciliacao
    
    def gerar_relatorio_divergencias(self, conciliacao):
        """Gera relatório detalhado de divergências"""
        
        # Resumo geral
        total_guias = len(conciliacao)
        valor_enviado_total = conciliacao['valor_enviado'].sum()
        valor_pago_total = conciliacao['valor_pago'].sum()
        diferenca_total = valor_pago_total - valor_enviado_total
        
        # Estatísticas por classificação
        stats_classificacao = conciliacao['classificacao'].value_counts()
        
        # Guias com divergência significativa (>5% ou valor absoluto >50)
        divergencias_significativas = conciliacao[
            (abs(conciliacao['percentual_divergencia']) > 5) | 
            (abs(conciliacao['diferenca_valor']) > 50)
        ]
        
        # Top 10 maiores divergências
        maiores_divergencias = conciliacao[
            conciliacao['diferenca_valor'] != 0
        ].nlargest(10, 'diferenca_valor')
        
        # Guias não processadas há mais de 30 dias
        nao_processadas_antigas = conciliacao[
            (conciliacao['classificacao'] == '🔴 Não Processado') & 
            (conciliacao['dias_em_aberto'] > 30)
        ]
        
        relatorio = {
            'resumo_geral': {
                'total_guias': total_guias,
                'valor_enviado_total': valor_enviado_total,
                'valor_pago_total': valor_pago_total,
                'diferenca_total': diferenca_total,
                'percentual_diferenca_total': (diferenca_total / valor_enviado_total * 100) if valor_enviado_total > 0 else 0
            },
            'stats_classificacao': stats_classificacao,
            'divergencias_significativas': divergencias_significativas,
            'maiores_divergencias': maiores_divergencias,
            'nao_processadas_antigas': nao_processadas_antigas
        }
        
        return relatorio
    
    def gerar_graficos_conciliacao(self, conciliacao, relatorio):
        """Gera gráficos da conciliação"""
        
        plt.style.use('default')
        fig, axes = plt.subplots(2, 2, figsize=(16, 12))
        
        # Gráfico 1: Distribuição por classificação
        classificacoes = relatorio['stats_classificacao']
        colors = ['#2ecc71', '#f39c12', '#e74c3c', '#9b59b6', '#34495e']
        axes[0,0].pie(classificacoes.values, labels=classificacoes.index, autopct='%1.1f%%', 
                     colors=colors[:len(classificacoes)])
        axes[0,0].set_title('Distribuição das Guias por Status', fontsize=14, fontweight='bold')
        
        # Gráfico 2: Divergências por convênio
        div_convenio = conciliacao.groupby('convenio').agg({
            'diferenca_valor': 'sum',
            'numero_guia': 'count'
        }).sort_values('diferenca_valor')
        
        bars = axes[0,1].bar(div_convenio.index, div_convenio['diferenca_valor'], 
                            color=['red' if x < 0 else 'green' for x in div_convenio['diferenca_valor']])
        axes[0,1].set_title('Divergência Total por Convênio (R$)', fontsize=14, fontweight='bold')
        axes[0,1].tick_params(axis='x', rotation=45)
        axes[0,1].axhline(y=0, color='black', linestyle='-', linewidth=0.5)
        
        # Adicionar valores nas barras
        for bar in bars:
            height = bar.get_height()
            axes[0,1].annotate(f'R$ {height:,.0f}',
                              xy=(bar.get_x() + bar.get_width() / 2, height),
                              xytext=(0, 3 if height >= 0 else -15),
                              textcoords="offset points",
                              ha='center', va='bottom' if height >= 0 else 'top',
                              fontsize=9)
        
        # Gráfico 3: Evolução temporal das divergências
        conciliacao_copy = conciliacao.copy()
        conciliacao_copy['mes_envio'] = conciliacao_copy['data_envio'].dt.to_period('M')
        evolucao = conciliacao_copy.groupby('mes_envio').agg({
            'diferenca_valor': 'sum',
            'numero_guia': 'count'
        })
        
        axes[1,0].plot(range(len(evolucao)), evolucao['diferenca_valor'], marker='o', linewidth=2, color='#e74c3c')
        axes[1,0].set_title('Evolução das Divergências por Mês', fontsize=14, fontweight='bold')
        axes[1,0].set_xlabel('Mês')
        axes[1,0].set_ylabel('Divergência Total (R$)')
        axes[1,0].tick_params(axis='x', rotation=45)
        axes[1,0].grid(True, alpha=0.3)
        
        # Gráfico 4: Distribuição de valores divergentes
        divergencias_valores = conciliacao[conciliacao['diferenca_valor'] != 0]['diferenca_valor']
        axes[1,1].hist(divergencias_valores, bins=20, edgecolor='black', alpha=0.7, color='#3498db')
        axes[1,1].set_title('Distribuição dos Valores Divergentes', fontsize=14, fontweight='bold')
        axes[1,1].set_xlabel('Valor da Divergência (R$)')
        axes[1,1].set_ylabel('Quantidade de Guias')
        axes[1,1].axvline(x=0, color='red', linestyle='--', alpha=0.7)
        
        plt.tight_layout()
        plt.savefig('graficos/conciliacao_guias.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        print("✅ Gráficos gerados: graficos/conciliacao_guias.png")
    
    def gerar_excel_conciliacao(self, conciliacao, relatorio):
        """Gera arquivo Excel completo com a conciliação"""
        
        filename = f"conciliacao/Conciliacao_Guias_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            
            # Aba 1: Resumo Executivo
            resumo_data = [
                ['Indicador', 'Valor'],
                ['Total de Guias Analisadas', f"{relatorio['resumo_geral']['total_guias']:,}"],
                ['Valor Total Enviado', f"R$ {relatorio['resumo_geral']['valor_enviado_total']:,.2f}"],
                ['Valor Total Pago', f"R$ {relatorio['resumo_geral']['valor_pago_total']:,.2f}"],
                ['Diferença Total', f"R$ {relatorio['resumo_geral']['diferenca_total']:,.2f}"],
                ['% Diferença Total', f"{relatorio['resumo_geral']['percentual_diferenca_total']:.2f}%"],
                ['', ''],
                ['Status das Guias', 'Quantidade'],
            ]
            
            for status, qtd in relatorio['stats_classificacao'].items():
                resumo_data.append([status, qtd])
            
            df_resumo = pd.DataFrame(resumo_data[1:], columns=resumo_data[0])
            df_resumo.to_excel(writer, sheet_name='Resumo Executivo', index=False)
            
            # Aba 2: Conciliação Completa
            colunas_conciliacao = [
                'numero_guia', 'convenio', 'data_atendimento', 'paciente', 'procedimento',
                'valor_enviado', 'valor_pago', 'diferenca_valor', 'percentual_divergencia',
                'classificacao', 'status_operadora', 'motivo_glosa', 'data_envio', 'dias_em_aberto'
            ]
            
            conciliacao_export = conciliacao[colunas_conciliacao].copy()
            conciliacao_export = conciliacao_export.sort_values(['classificacao', 'diferenca_valor'], ascending=[True, False])
            conciliacao_export.to_excel(writer, sheet_name='Conciliação Completa', index=False)
            
            # Aba 3: Divergências Significativas
            if not relatorio['divergencias_significativas'].empty:
                divergencias_export = relatorio['divergencias_significativas'][colunas_conciliacao]
                divergencias_export.to_excel(writer, sheet_name='Divergências Significativas', index=False)
            
            # Aba 4: Top 10 Maiores Divergências
            if not relatorio['maiores_divergencias'].empty:
                maiores_export = relatorio['maiores_divergencias'][colunas_conciliacao]
                maiores_export.to_excel(writer, sheet_name='Maiores Divergências', index=False)
            
            # Aba 5: Não Processadas (>30 dias)
            if not relatorio['nao_processadas_antigas'].empty:
                nao_proc_export = relatorio['nao_processadas_antigas'][colunas_conciliacao]
                nao_proc_export.to_excel(writer, sheet_name='Não Processadas +30d', index=False)
            
            # Aba 6: Resumo por Convênio
            resumo_convenio = conciliacao.groupby('convenio').agg({
                'numero_guia': 'count',
                'valor_enviado': 'sum',
                'valor_pago': 'sum',
                'diferenca_valor': 'sum'
            }).round(2)
            resumo_convenio['percentual_diferenca'] = (
                (resumo_convenio['diferenca_valor'] / resumo_convenio['valor_enviado']) * 100
            ).round(2)
            resumo_convenio.to_excel(writer, sheet_name='Resumo por Convênio')
        
        # Formatação do Excel
        self.formatar_excel(filename)
        
        print(f"✅ Relatório Excel gerado: {filename}")
        return filename
    
    def formatar_excel(self, filename):
        """Aplica formatação ao arquivo Excel"""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = load_workbook(filename)
        
        # Formatação geral para todas as abas
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Cabeçalhos
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
    
    def executar_conciliacao_completa(self):
        """Executa todo o processo de conciliação"""
        try:
            print("🔄 Iniciando Conciliação de Guias...")
            print("="*50)
            
            # 1. Carregar dados
            print("📊 Carregando dados...")
            df_interno, df_operadora = self.carregar_dados()
            
            # 2. Realizar conciliação
            print("🔍 Realizando conciliação...")
            conciliacao = self.conciliar_guias(df_interno, df_operadora)
            
            # 3. Gerar relatório
            print("📈 Gerando análises...")
            relatorio = self.gerar_relatorio_divergencias(conciliacao)
            
            # 4. Gerar gráficos
            print("📊 Criando gráficos...")
            self.gerar_graficos_conciliacao(conciliacao, relatorio)
            
            # 5. Gerar Excel
            print("📋 Gerando relatório Excel...")
            arquivo_excel = self.gerar_excel_conciliacao(conciliacao, relatorio)
            
            # 6. Resumo final
            print("\n" + "="*60)
            print("📊 RESULTADO DA CONCILIAÇÃO")
            print("="*60)
            
            resumo = relatorio['resumo_geral']
            stats = relatorio['stats_classificacao']
            
            print(f"📋 Total de Guias Analisadas: {resumo['total_guias']:,}")
            print(f"💰 Valor Enviado: R$ {resumo['valor_enviado_total']:,.2f}")
            print(f"💳 Valor Pago: R$ {resumo['valor_pago_total']:,.2f}")
            print(f"📊 Diferença: R$ {resumo['diferenca_total']:,.2f} ({resumo['percentual_diferenca_total']:.2f}%)")
            
            print("\n📈 DISTRIBUIÇÃO DAS GUIAS:")
            for status, qtd in stats.items():
                percentual = (qtd / resumo['total_guias']) * 100
                print(f"   {status}: {qtd:,} ({percentual:.1f}%)")
            
            print(f"\n⚠️ Divergências Significativas: {len(relatorio['divergencias_significativas']):,}")
            print(f"🕐 Não Processadas +30 dias: {len(relatorio['nao_processadas_antigas']):,}")
            
            print(f"\n📁 ARQUIVOS GERADOS:")
            print(f"   • {arquivo_excel}")
            print(f"   • graficos/conciliacao_guias.png")
            
            print("\n✅ CONCILIAÇÃO CONCLUÍDA COM SUCESSO!")
            
            # Alertas importantes
            if resumo['percentual_diferenca_total'] < -5:
                print("\n🚨 ALERTA: Diferença significativa - Hospital recebendo MENOS que enviado!")
            elif resumo['percentual_diferenca_total'] > 5:
                print("\n✨ Hospital recebendo MAIS que enviado - Verificar se está correto")
            
            if len(relatorio['nao_processadas_antigas']) > 0:
                print(f"\n⏰ AÇÃO NECESSÁRIA: {len(relatorio['nao_processadas_antigas'])} guias não processadas há mais de 30 dias!")
            
        except Exception as e:
            print(f"❌ Erro na conciliação: {str(e)}")
            import traceback
            traceback.print_exc()

# EXECUÇÃO PRINCIPAL
if __name__ == "__main__":
    conciliador = ConciliadorGuias()
    conciliador.executar_conciliacao_completa()
    
    print("\n" + "="*60)
    print("🎯 COMO USAR ESTE PROJETO NO SEU PORTFÓLIO:")
    print("="*60)
    print("1. 📊 Substitua dados fictícios por arquivos reais do hospital")
    print("2. 🔧 Configure para diferentes formatos de demonstrativos")
    print("3. 📧 Adicione envio automático para gestores")
    print("4. 🕒 Agende execução diária/semanal")
    print("5. 📱 Crie alertas para divergências críticas")
    print("6. 🌐 Desenvolva dashboard web interativo")
    print("\n💡 Este projeto resolve um problema REAL e custoso dos hospitais!")
    print("🚀 Mostra domínio técnico E conhecimento específico do negócio!")